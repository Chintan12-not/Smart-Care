import openpyxl
import json
import os

excel_path = r"c:\Users\lenovo\Documents\Smart Care\phone data\phone_models_10_brands.xlsx"
wb = openpyxl.load_workbook(excel_path)

data = {}
print("Sheet names:", wb.sheetnames)

for sheet in wb.sheetnames:
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if rows:
        header = rows[0]
        data[sheet] = [dict(zip(header, r)) for r in rows[1:] if any(r)]

out_path = r"c:\Users\lenovo\Documents\Smart Care\scratch\parsed_phone_data.json"
os.makedirs(os.path.dirname(out_path), exist_ok=True)
with open(out_path, "w", encoding="utf-8") as f:
    json.dump(data, f, indent=2, ensure_ascii=False)

print(f"Successfully exported data to {out_path}")
